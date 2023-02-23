from selenium import webdriver
from selenium.webdriver.common.by import By
#from selenium.webdriver.chrome.options import Options
from json import loads,dumps
import time
from openpyxl import Workbook,load_workbook
import os
from selenium.webdriver.common.keys import Keys

url = "https://meet.google.com/dke-mjah-gnw"
config = os.path.join(os.path.dirname(os.path.abspath(__file__)),'config')

class Functions():
    def __init__(self) -> None:
            answer = input('請選擇功能:\n(1 離開\n')
            if answer == '1':
                self.quit()
    def quit(self):
        roll_call_list.save_file()
        driver.quit()
        quit()

class Member_events():
    members = []
    def __init__(self)->None:
        while True:
            self.detection()
    def detection(self)->None:
        elements_for_now = driver.find_elements(By.CLASS_NAME,"zWGUib")
        members_for_now=[]
        for member in elements_for_now:
            try:
                members_for_now.append(member.text)
            except:
                pass
        if members_for_now != self.members:
            if len(members_for_now) > len(self.members):
                self.member_joined(members_for_now)
            else:
                self.member_left(members_for_now)
            self.members = members_for_now
            roll_call_list.save_file()
    def member_joined(self,members_for_now)->None:
        for member in members_for_now:
            if not member in self.members:
                print(member+'加入了!')
                roll_call_list.join(member)
                
    def member_left(self,members_for_now)->None:
        for member in self.members:
            if not member in members_for_now:
                print(member+'離開了!')
                roll_call_list.left(member)
                

class Roll_call_list():
    __file_name = ''
    def __init__(self):
        self.__file_name = os.path.join(os.path.dirname(os.path.abspath(__file__)),input("點名單檔案名稱:")+'.xlsx')

        if os.path.isfile(self.__file_name):
            self.load()
        else:
            self.new_workbook()
        self.save_file()
    def new_workbook(self):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.paint_sheet()
    def load(self):
        self.workbook = load_workbook(self.__file_name)
        self.worksheet = self.workbook.create_sheet()
        self.paint_sheet()
    def paint_sheet(self):
        local_time = time.localtime(time.time())
        self.worksheet.title = f'{local_time.tm_hour}-{local_time.tm_min}'
        self.worksheet.merge_cells('A1:B1')
        self.worksheet['A1']='加入'
        self.worksheet.merge_cells('C1:D1')
        self.worksheet['C1']='離開'
        self.worksheet['E1']='點名紀錄'
        self.worksheet.column_dimensions['A'].width = 25
        self.worksheet.column_dimensions['B'].width = 20
        self.worksheet.column_dimensions['C'].width = 25
        self.worksheet.column_dimensions['D'].width = 20
    def save_file(self):
        while True:
            try:
                self.workbook.save(self.__file_name)
            except PermissionError:
                input('無法儲存檔案，請關閉檔案後按enter重試')
            except Exception as exception:
                print('儲存檔案時發生錯誤:'+str(exception))
                input('按Enter重試')
            else:
                print('檔案儲存成功!')
                return
    def join(self,user_name:str):
        current_cell = 2
        
        while True:
            if self.worksheet['A'+str(current_cell)].value:
                current_cell+=1
                continue
            
            self.worksheet['A'+str(current_cell)] = time.ctime()
            self.worksheet['B'+str(current_cell)] = user_name
            
            return
    def left(self,user_name:str):
        current_cell = 2
        while True:
            if self.worksheet['C'+str(current_cell)].value:
                current_cell+=1
                continue
            self.worksheet['C'+str(current_cell)] = time.ctime()
            self.worksheet['D'+str(current_cell)] = user_name
            
            return

class Login_Google():
    def __init__(self) -> None:
        driver.get(url)
        if os.path.isfile('config'):
            #add cookie
            try:
                self.login_with_cookie()
            except:
                print('登入憑證無效，重新建立中...')
                self.manual_login()
        else:
            print('未偵測到登入憑證，請手動登入')
            self.manual_login()
        driver.implicitly_wait(10)
        driver.get(url)
    def login_with_cookie(self):
        driver.delete_all_cookies()
        with open(config,"r") as file:
            cookie = loads(file.read())
            for name,value in cookie.items():
                driver.add_cookie({"name":name,"value":value})
    def manual_login(self):
        driver.delete_all_cookies()
        driver.get("https://accounts.google.com/v3/signin/identifier?dsh=S-905868804%3A1676301694738053&_ga=2.207397329.35016544.1676301693-561019711.1676301693&continue=https%3A%2F%2Fmeet.google.com%3Fhs%3D193&ltmpl=meet&o_ref=https%3A%2F%2Fmeet.google.com%2F&flowName=GlifWebSignIn&flowEntry=ServiceLogin&ifkv=AWnogHdWVd-qRecSxEULl_n6vHYB07-UsP6Iqy7uBGNwMSXZ3GGXmR3rUzCtiuwX1w4tF4PMLDRIDQ")

        driver.implicitly_wait(30)

        # Enter your username and password
        
        username = driver.find_element(By.NAME,"identifier")
        username.send_keys(input('輸入您的gmail帳號:'))
        username.send_keys(Keys.ENTER)

        # password
        driver.implicitly_wait(10)

        passwd = driver.find_element(By.NAME,"Passwd")
        passwd.send_keys(input('輸入您的密碼:'))

        # Click the next button
        passwd.send_keys(Keys.ENTER)
        time.sleep(3)
        web_cookies = driver.get_cookies()
        cookies = {}
        for cookie in web_cookies:
            cookies[cookie['name']]=cookie['value']
        with open(config,'x')as file:
            file.write(dumps(cookies))
        self.login_with_cookie()


#preparation
print('初始化軟體...')
print('建立點名單...')
roll_call_list = Roll_call_list()
print('準備加入會議...')
#options.add_argument('--headless')
#options.add_argument('--disable-gpu')
driver=webdriver.Edge()
driver.maximize_window()

print('嘗試登入...')
while True:
    Login_Google()

    driver.implicitly_wait(4)
    try:
        driver.find_element(By.CLASS_NAME,"AjXHhf").click()
    except:
        print('登入失敗')
        os.remove('config')
        continue
    else: 
        link = driver.find_element(By.CLASS_NAME,"QJgqC")
        link.click()
        driver.implicitly_wait(30)
        driver.find_elements(By.CLASS_NAME,"boDUxc")[1].click()
        break
print('會議加入成功!')
time.sleep(1)
os.system('cls')


Member_events()
#driver.close()